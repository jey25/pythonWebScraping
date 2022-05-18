import openpyxl
from requests import patch

# 엑셀 불러오기
path = r'C:\Users\jey25\pythonWebScraping\pythonAccel\참가자_data.xlsx'

wb = openpyxl.load_workbook(path)

# 엑셀 시트 선택
ws = wb['오징어게임']

# 데이터 수정
ws['A3'] = 456
ws['B3'] = '성기훈'

# 엑셀 저장하기
wb.save(path)